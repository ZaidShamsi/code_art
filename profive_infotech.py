import openpyxl
from openpyxl.styles import Alignment
import time

# Using selenium for data extraction
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# for suppressing the browser
from selenium.webdriver.firefox.options import Options

option = webdriver.FirefoxOptions()
option.add_argument('--headless')
s = Service(r"C:\Users\Alig\Desktop\webdrivers\geckodriver.exe")
#driver_object = webdriver.Firefox(service = s, options = option)
driver_object = webdriver.Firefox(service = s)
driver_object.implicitly_wait(10)

base_URL = "https://www.google.com/search?q=supplier+in+"
location = ['ahmedabad', 'bangalore', 'Chandigarh']
#location = ['ahmedabad']

URL = base_URL+location[0]

def extractText(URL):
    driver_object.get(URL)
    try:
        more_places_btn = driver_object.find_element(By.CSS_SELECTOR, "div[class = 'mSBLIb ndElDd'] a")
        more_places_btn.click()
    except:
        view_all_btn = driver_object.find_element(By.CSS_SELECTOR, "div[class = 'ndElDd'] a")
        view_all_btn.click()
    supplier_link_list = driver_object.find_elements(By.CSS_SELECTOR, "div[class = 'VkpGBb'] a[class = 'C8TUKc rllt__link a-no-hover-decoration']")
    address_list = []
    phone_list = []
    company_name_list = []
    email_id_list = []
    index = 0
    for supplier_link in supplier_link_list:
        supplier_link.click()
        if index == 0:
            print(index)
            company_name_1 = driver_object.find_element(By.CSS_SELECTOR, "div[class='SPZz6b'] h2 span")
            company_name_1 = company_name_1.get_attribute('innerText')
            company_name_list.append(company_name_1)
            try:
                address = driver_object.find_element(By.CSS_SELECTOR, "span[class='LrzXr']")
                address = address.get_attribute('innerText')
                address_list.append(address)
                phone = driver_object.find_element(By.CSS_SELECTOR, "span[class='LrzXr zdqRlf kno-fv'] a span")
                phone = phone.get_attribute('innerText')
                phone_list.append(phone)
            except:
                address_list.append('Not Found')
                phone_list.append('Not Found')
            try:
                website_btn = driver_object.find_element(By.CSS_SELECTOR, "div[class = 'QqG1Sd']")
                website_btn = website_btn.find_element(By.LINK_TEXT, 'Website')
                website_btn.click()
                contant_us_element = driver_object.find_element(By.LINK_TEXT, 'Contact Us')
                contant_us_element.click()
                break
            except:
                print('Website not available')
                email_id_list.append('Not Found')
        elif index > 0:
            print(index)
            company_name_2 = driver_object.find_element(By.CSS_SELECTOR, "div[class='SPZz6b'] h2 span")
            company_name_2 = company_name_2.get_attribute('innerText')
            while company_name_2 == company_name_1:
                print('inside while loop')
                time.sleep(2)
                company_name_2 = driver_object.find_element(By.CSS_SELECTOR, "div[class='SPZz6b'] h2 span")
                company_name_2 = company_name_2.get_attribute('innerText')
            print('outside while loop')
            company_name_list.append(company_name_2)
            company_name_1 = company_name_2
            try:
                address = driver_object.find_element(By.CSS_SELECTOR, "span[class='LrzXr']")
                address = address.get_attribute('innerText')
                address_list.append(address)
                phone = driver_object.find_element(By.CSS_SELECTOR, "span[class='LrzXr zdqRlf kno-fv'] a span")
                phone = phone.get_attribute('innerText')
                phone_list.append(phone)
            except:
                address_list.append('Not Found')
                phone_list.append('Not Found')
            try:
                website_btn = driver_object.find_element(By.CSS_SELECTOR, "div[class = 'QqG1Sd']")
                website_btn = website_btn.find_element(By.LINK_TEXT, 'Website')
                website_btn.click()
                contant_us_element = driver_object.find_element(By.LINK_TEXT, 'Contact Us')
                contant_us_element.click()
                break
            except:
                print('Website not available')
                email_id_list.append('Not Found')
        index += 1
    return company_name_list, address_list, phone_list

#extractText(URL)
result = extractText(URL)

#filepath = r'C:\Users\Alig\Downloads\Output Format.xlsx'
'''
wb = openpyxl.load_workbook(filepath)
sh = wb['Sheet1']
sh.column_dimensions['B'].width = 30
sh.column_dimensions['C'].width = 80

for index_location in range(0, len(location)):
    result = extractText(base_URL+location[index_location])
    for index_result in range(0, len(result[0])):
        for col in range(1, sh.max_column):
            if col == 1:
                sh.cell(row = ((20*index_location) + (index_result + 2)), column = col).value = ((20*index_location) + (index_result + 1))
            else:
                print(result[col-2][index_result])
                sh.cell(row = ((20*index_location) + (index_result + 2)), column = col).value = result[col-2][index_result]
                sh.cell(row = ((20*index_location) + (index_result + 2)), column = col).alignment = Alignment(wrap_text = True)
wb.save(filepath)
#'''

#driver_object.close()

#'''
print(result[0])
print(result[1])
print(result[2])
#'''
